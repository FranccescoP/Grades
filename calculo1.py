import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.drawing.image import Image
import json
import os

# Leer el archivo JSON
with open('alumnos_datos.json', 'r', encoding='utf-8') as f:
    datos_json = json.load(f)

# Función para calcular el promedio final
def calcular_promedio(evaluaciones, examen_final):
    return (sum(evaluaciones) * 0.1 + examen_final * 0.6)

# Función para determinar el estado (Aprobado o Desaprobado)
def obtener_estado(promedio_final):
    return "Aprobado" if promedio_final >= 10 else "Desaprobado"

# Crear un libro de trabajo (workbook)
wb = openpyxl.Workbook()
# Seleccionar la hoja activa
sheet = wb.active
sheet.title = "Notas de Alumnos"

# Añadir el encabezado
sheet['A1'] = "UNIVERSIDAD NACIONAL DE LA AMAZONIA PERUANA"
sheet['A2'] = "ESCUELA DE POSTGRADO"
sheet['A3'] = "MAESTRIA EN EDUCACION CON MENCION EN GESTION EMPRESARIAL -2025-"
sheet['B4'] = "ASIGNATURA"
sheet['B5'] = "DOCENTE"
sheet['B6'] = "FECHA"
sheet['E4'] = ":MDU-101 Calculo I"
sheet['E5'] = ":Dr. Frank Suarez"
sheet['E6'] = ":Del 01 de Enero al 30 de Marzo del 2025" 
sheet['C7'] = "Control de notas" 

# Estilo para el encabezado
encabezado_font = Font(size=14, bold=True)
sheet['A1'].font = encabezado_font
sheet['A2'].font = encabezado_font
sheet['A3'].font = encabezado_font
sheet['B4'].font = encabezado_font
sheet['B5'].font = encabezado_font
sheet['B6'].font = encabezado_font
sheet['E4'].font = encabezado_font
sheet['E5'].font = encabezado_font
sheet['E6'].font = encabezado_font
sheet['C7'].font = encabezado_font

# Justificar el encabezado
justificacion = Alignment(horizontal="center", vertical="center")
sheet['A1'].alignment = justificacion
sheet['A2'].alignment = justificacion
sheet['A3'].alignment = justificacion
justificacion = Alignment(horizontal="left", vertical="center")
sheet['B4'].alignment = justificacion
sheet['B5'].alignment = justificacion
sheet['B6'].alignment = justificacion
justificacion = Alignment(horizontal="center", vertical="center")
sheet['C7'].alignment = justificacion

# Ajustar el tamaño de las celdas para el encabezado
sheet.merge_cells('A1:H1')
sheet.merge_cells('A2:H2')
sheet.merge_cells('A3:H3')
sheet.merge_cells('B4:D4')
sheet.merge_cells('B5:D5')
sheet.merge_cells('B6:D6')
sheet.merge_cells('E4:H4')
sheet.merge_cells('E5:H5')
sheet.merge_cells('E6:H6')
sheet.merge_cells('C7:H7')

# Títulos de las columnas
sheet['A8'] = "N°"
sheet['B8'] = "Apellido y nombre"
sheet['C8'] = "Evaluación 1"
sheet['D8'] = "Evaluación 2"
sheet['E8'] = "Evaluación 3"
sheet['F8'] = "Evaluación 4"
sheet['G8'] = "Examen Final"
sheet['H8'] = "Promedio Final"
sheet['I8'] = "Estado"

# Ajustar el tamaño de las columnas
sheet.column_dimensions['A'].width = 5  # N°
sheet.column_dimensions['B'].width = 45  # Apellido y nombre
sheet.column_dimensions['C'].width = 13  # Evaluacion 1
sheet.column_dimensions['D'].width = 13  # Evaluación 2
sheet.column_dimensions['E'].width = 13  # Evaluación 3
sheet.column_dimensions['F'].width = 13  # Evaluación 4
sheet.column_dimensions['G'].width = 13  # Examen Final
sheet.column_dimensions['H'].width = 15  # Promedio Final
sheet.column_dimensions['I'].width = 20  # Estado

# Justificar los títulos de las columnas
for col in ['A8', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8']:
    sheet[col].alignment = justificacion

# Ajustar la altura de las filas (modificada para mayor espacio)
sheet.row_dimensions[8].height = 15  # Aseguramos que la fila 8 tenga suficiente altura
sheet.row_dimensions[9].height = 15  # Ajustar la altura de las filas normales (a partir de la fila 9)

# Insertar los números del 1 al 25 en la columna A, debajo de la celda A8
for i in range(1, 26):  # Desde 1 hasta 25
    sheet[f"A{i+8}"] = i  # Colocamos los números en las filas 9 a 33

# Insertar los datos de los alumnos en las filas correspondientes
for i, alumno in enumerate(datos_json["alumnos"], start=9):  # Comenzar desde la fila 9
    nombre = alumno["nombre"]
    evaluaciones = alumno["evaluaciones"]
    examen_final = alumno["examen_final"]
    
    promedio_final = calcular_promedio(evaluaciones, examen_final)
    estado = obtener_estado(promedio_final)
    
    # Asignar los valores a las celdas
    sheet[f"B{i}"] = nombre
    sheet[f"C{i}"] = evaluaciones[0]
    sheet[f"D{i}"] = evaluaciones[1]
    sheet[f"E{i}"] = evaluaciones[2]
    sheet[f"F{i}"] = evaluaciones[3]
    sheet[f"G{i}"] = examen_final
    sheet[f"H{i}"] = round(promedio_final, 2)
    sheet[f"I{i}"] = estado

# Definir un borde fino para las celdas
border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                top=Side(border_style="thin"), bottom=Side(border_style="thin"))

# Aplicar bordes a todas las celdas con datos (comenzando desde la fila 9)
for i in range(8, 34):  # Desde la fila 9 hasta la fila 33
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        sheet[f"{col}{i}"].border = border

# Justificar las celdas de cada alumno
for i in range(8, 34):  # Desde la fila 9 hasta la fila 33
    for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        sheet[f"{col}{i}"].alignment = justificacion

# Justificar la columna B hacia la izquierda (de B8 a B33)
for i in range(8, 34):  # Desde la fila 8 hasta la fila 33
    sheet[f"B{i}"].alignment = Alignment(horizontal="left", vertical="center")

# Insertar el logo de la universidad en la celda I1
logo_path = "20.png"  # Ruta del archivo de imagen (asegúrate de que esté en el mismo directorio)
logo = Image(logo_path)

# Ajustar el tamaño de la imagen para que se ajuste a las celdas I1-I6
# Aquí se ajustan tanto la altura de las filas como la columna
sheet.row_dimensions[1].height = 20  # Ajustar la altura de las filas 1 a 6
sheet.row_dimensions[2].height = 20
sheet.row_dimensions[3].height = 20
sheet.row_dimensions[4].height = 20
sheet.row_dimensions[5].height = 20
sheet.row_dimensions[6].height = 20

sheet.column_dimensions['I'].width = 25  # Ajustar el tamaño de la columna I para la imagen

# Redimensionar la imagen para que se ajuste a las celdas
logo.width = 210  # Aumentar un poco el ancho de la imagen (en píxeles)
logo.height = 170  # Ajustar la altura de la imagen (en píxeles)

# Colocar la imagen en I1
sheet.add_image(logo, "I1")

# Añadir el pie de página en la última fila (puedes colocar en la fila 34, por ejemplo)
sheet['A34'] = "Escuela de Postgrado UNAP"
sheet['A34'].font = Font(size=12, italic=True)
sheet['A34'].alignment = Alignment(horizontal="center", vertical="center")
sheet.merge_cells('A34:I34')

# Intentar guardar el archivo con manejo de permisos
archivo_guardado = False
intentos = 0
while not archivo_guardado and intentos < 3:
    try:
        # Verificar si el archivo ya existe y eliminarlo si es necesario
        if os.path.exists("alumnosdatos1.xlsx"):
            os.remove("alumnosdatos1.xlsx")
        
        wb.save("alumnosdatos1.xlsx")
        archivo_guardado = True
        print("Archivo guardado exitosamente.")
    except PermissionError:
        print("Error de permisos. Intentando nuevamente...")
        intentos += 1

    if intentos == 3:
        print("No se pudo guardar el archivo después de varios intentos.")
